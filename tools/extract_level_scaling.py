"""캐릭터 레벨 스케일링 자료를 엑셀에서 뽑아 CSV로 굽는다.

    python tools/extract_level_scaling.py [엑셀경로]

원본(위키에서 받아 적은 엑셀)은 저장소 밖에서도 바뀔 수 있고 openpyxl 없이는 못 읽는다.
엔진이 실제로 읽는 것은 여기서 만든 gidc/core/data/character_level_scaling/*.csv 넷뿐이고,
이 스크립트는 그 표가 어디서 어떻게 나왔는지를 남기는 역할이다 — 숫자가 의심스러우면
다시 돌리면 된다.

표 하나가 파일 하나다. 넷 다 순수 CSV라 엑셀·깃 diff·눈으로 그대로 읽힌다
(주석 줄을 넣지 않는 것도 그래서다 — 설명은 core/base_stats.py 에 있다).

    스탯 = 기초값 × 레벨 배율 + 최대 어센션 값 × 어센션 배율

엑셀 시트 넷 중 셋을 쓴다.
  · Base Value          캐릭터 → 기초 HP/공격력/방어력 (Lv.1)
  · Max Ascension Values 캐릭터 → 6돌파에서 더해지는 몫
  · Level Scaling        레벨 → 4성/5성 배율 (A~C열의 Lv.1~90만 쓴다. F~I열의 95·100은
                         캐릭터 상한을 넘고 HP/DEF와 ATK가 갈려 있어 다른 용도로 보인다)
  · Ascension Multiplier 돌파 단계 → 배율. 아래 _ascension_numerators 참고.
"""
from __future__ import annotations

import csv
import pathlib
import sys

import openpyxl

ROOT   = pathlib.Path(__file__).resolve().parent.parent
XLSX   = ROOT / "Genshin Imapct Character Level Scaling.xlsx"
OUTDIR = ROOT / "gidc" / "core" / "data" / "character_level_scaling"

MAX_LEVEL = 90
RARITIES  = (4, 5)


def _ascension_numerators(row: tuple) -> tuple[list[int], int]:
    """돌파 배율 행을 분자 목록과 분모로 되돌린다.

    위키는 이 배율을 0/182, 38/182 … 182/182 분수로 적는데, 엑셀로 옮겨지면서 슬래시가
    사라져 182, 38182, 65182 … 182182 라는 숫자로 남았다. 뒤 세 자리가 분모다.

    되돌린 값이 맞는지는 이 파일 밖에서 확인했다 — 이 표로 계산한 Lv.90 6돌파 스탯이
    엔진에 이미 박혀 있던 값과 9명 중 8명에서 ±1 안에 들어온다(어긋난 실로닌은 코드 쪽이
    Lv.81 값이었다). 그래도 아래 가정 검사로 형태만은 매번 확인한다."""
    raw = [int(v) for v in row[1:] if v is not None]
    if len(raw) != 7:
        raise ValueError(f"돌파 단계는 0~6, 즉 7칸이어야 합니다. (읽은 값: {raw})")

    denominator = int(str(raw[-1])[-3:])
    numerators  = [int(str(v)[:-3] or 0) for v in raw]

    if numerators[0] != 0 or numerators[-1] != denominator:
        raise ValueError(f"0돌파는 0, 6돌파는 분모와 같아야 합니다. {numerators}/{denominator}")
    if any(a >= b for a, b in zip(numerators, numerators[1:])):
        raise ValueError(f"돌파 배율은 단계마다 커져야 합니다. {numerators}")
    return numerators, denominator


def _character_stats(ws) -> tuple[dict[str, dict[str, float]], list[str]]:
    """캐릭터 → 스탯 셋. 값이 아직 안 채워진 행(위키가 '?'로 둔 캐릭터)은 걸러 따로 돌려준다."""
    out, skipped = {}, []
    for name, hp, atk, defense in ws.iter_rows(min_row=2, values_only=True):
        if name is None:
            continue
        if name in out or name in skipped:
            raise ValueError(f"'{ws.title}' 시트에 '{name}'이(가) 두 번 나옵니다.")
        if not all(isinstance(v, (int, float)) for v in (hp, atk, defense)):
            skipped.append(name)
            continue
        out[name] = {"hp": float(hp), "atk": float(atk), "def": float(defense)}
    return out, skipped


def _level_multipliers(ws) -> dict[int, dict[int, float]]:
    """A~C열: 레벨 → 4성/5성 배율."""
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        level, four, five = row[0], row[1], row[2]
        if level is None:
            continue
        out[int(level)] = {4: float(four), 5: float(five)}

    missing = [lv for lv in range(1, MAX_LEVEL + 1) if lv not in out]
    if missing:
        raise ValueError(f"레벨 배율 표에 빠진 레벨이 있습니다: {missing}")
    return out


def main() -> None:
    path = pathlib.Path(sys.argv[1]) if len(sys.argv) > 1 else XLSX
    if not path.exists():
        raise SystemExit(f"엑셀을 찾을 수 없습니다: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    base,      base_skipped = _character_stats(wb["Base Value"])
    ascension, asc_skipped  = _character_stats(wb["Max Ascension Values"])
    levels    = _level_multipliers(wb["Level Scaling"])
    numerators, denominator = _ascension_numerators(
        next(wb["Ascension Multiplier"].iter_rows(min_row=2, max_row=2, values_only=True))
    )

    if set(base) != set(ascension):
        only = set(base) ^ set(ascension)
        raise ValueError(f"두 캐릭터 시트의 명단이 다릅니다: {sorted(only)}")

    OUTDIR.mkdir(parents=True, exist_ok=True)
    written = [
        _write("ascension_multiplier.csv", ["phase", "numerator", "denominator"],
               [[phase, n, denominator] for phase, n in enumerate(numerators)]),
        _write("level_multiplier.csv", ["level", "4star", "5star"],
               [[level, *(levels[level][r] for r in RARITIES)]
                for level in sorted(levels) if level <= MAX_LEVEL]),
        _write("character_base_value.csv", ["character", "hp", "atk", "def"],
               [[name, *(base[name][k] for k in ("hp", "atk", "def"))] for name in sorted(base)]),
        _write("character_max_ascension.csv", ["character", "hp", "atk", "def"],
               [[name, *(ascension[name][k] for k in ("hp", "atk", "def"))]
                for name in sorted(ascension)]),
    ]

    print(f"{OUTDIR.relative_to(ROOT)}")
    for name, rows, size in written:
        print(f"  {name:28s} {rows:4d}행 · {size / 1024:5.1f} KB")
    print(f"  캐릭터 {len(base)}명 · 레벨 1~{MAX_LEVEL} · 돌파 배율 {numerators} / {denominator}")
    if skipped := sorted(set(base_skipped) | set(asc_skipped)):
        print(f"  값이 비어 있어 뺀 캐릭터 {len(skipped)}명: {', '.join(skipped)}")


def _write(name: str, header: list[str], rows: list[list]) -> tuple[str, int, int]:
    path = OUTDIR / name
    # newline="" 이 없으면 윈도우에서 줄바꿈이 \r\r\n 으로 겹친다(csv 모듈이 이미 \r\n 을 쓴다).
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)
    return name, len(rows), path.stat().st_size


if __name__ == "__main__":
    for _stream in (sys.stdout, sys.stderr):
        if hasattr(_stream, "reconfigure") and not _stream.isatty():
            _stream.reconfigure(encoding="utf-8")
    main()
